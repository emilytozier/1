import openpyxl
import math
wb = openpyxl.load_workbook('trekking3.xlsx')
#this part reads the "guide" names of products and corresponding calories,fats, carbohydrates and proteins
sheet = wb['Справочник']
product = dict()
for i in range(2, sheet.max_row+1):
    product[sheet.cell(row=i,column=1).value] = \
        list([sheet.cell(row=i,column=x).value
                      if sheet.cell(row=i,column=x).value != None
                      else 0
                      for x in range(2,sheet.max_column+1)])

#this part takes the certain names form 'distribution' list and matches them with mass in gramms (as vocabulary 'mday')
sheet = wb['Раскладка']
mday = dict()
for i in range(2, sheet.max_row+1):
    name = sheet.cell(row=i,column=2).value #takes name from distribution list
    massa = sheet.cell(row=i,column=3).value #takes mass of products, cosumes each day
    mday[name] = mday.get(name, 0) + massa
#this part takes the certain product in product (dict) and through v_data determine how much of the calories, fats....  is consumed every day for one product

f_data = [0,0,0,0]
j_data={}
for k,v in mday.items():
    k_data = product[k]
    v_data = [i * (v/100)/2 for i in k_data]
    f_data = [(f_data[i] + v_data[i]) for i in range(len(f_data))]
    #print(k, v_data, end='\n\n')  # ТЕСТ ДАННЫХ
    j_data[k]=v_data

#this part sum the values from f_data to determine how much calories, fat, ... is consumed every day in total
print(j_data)
sheet = wb['Раскладка']
#this little part is devoted to juxstapose for each day the set of products
days=dict()
for i in range(2, sheet.max_row+1):
    day = sheet.cell(row=i,column=1).value
    kal=v in j_data.items()
        days[day]= days.get(day, 0) + kal
print(days)




