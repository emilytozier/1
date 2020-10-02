import openpyxl
import operator

wb = openpyxl.load_workbook('trekking1.xlsx')
sh = wb.active
rows = sh.max_row
columns = sh.max_column
list_of_name_food = []
list_of_value_calory = []
final_dict = {}
for row in range(2, rows + 1):
    for col in range(1, 2):
        name_food = sh.cell(row, col).value
        list_of_name_food.append(name_food)

for row1 in range(2, rows + 1):
    for col1 in range(2, 3):
        value_calory = sh.cell(row1, col1).value
        list_of_value_calory.append(value_calory)

cnt = 0
for i in list_of_name_food:
    final_dict[i] = list_of_value_calory[cnt]
    cnt += 1

for k, v in sorted(sorted(final_dict.items(), key=operator.itemgetter(0)), key=operator.itemgetter(1), reverse=True):
    print(k)