import openpyxl
from openpyxl import load_workbook

wb=openpyxl.load_workbook('salaries.xlsx')

sheet=wb.get_sheet_by_name('Лист1')
k=[]
for i in range(2,sheet.max_row+1):
    for j in range(2,sheet.max_column+1):
        d=sheet.cell(row=i,column=j)
        k.append(d.value)
ch=sorted(len(k)//2)

print(ch)

#print(vals[len(vals)//2],sum(vals)/len(vals)) #медианная з/п, средняя з.п.
#print(sheet['A1'].value)