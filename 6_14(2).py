# -*- coding: utf-8 -*-
#в таблице зарплаты для разных профессий. выведите регион с самой высокой медианной зарплатой и название профессии с самой высокой средней зарплатой
import openpyxl
from statistics import median, mean

wb = openpyxl.load_workbook('salaries.xlsx')
sheet = wb.get_sheet_by_name('Лист1')

regions, proffs = {}, {}
for col in range(2, sheet.max_column + 1):
    for row in range(2, sheet.max_row + 1):
        salary = sheet.cell(row=row, column=col).value
        reg = sheet.cell(row=row, column=1).value
        prof = sheet.cell(row=1, column=col).value

        if reg not in regions:
            regions[reg] = []
        regions[reg].append(salary)

        if prof not in proffs:
            proffs[prof] = []
        proffs[prof].append(salary)

for reg in regions:
    regions[reg] = mean(regions[reg])
for prof in proffs:
    proffs[prof] = mean(proffs[prof])

print(sorted(regions.items(), key=lambda x: x[1])[-1][0],
sorted(proffs.items(), key=lambda x: x[1])[-1][0], sep=' ')

  
