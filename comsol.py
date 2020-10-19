import openpyxl
from openpyxl import load_workbook
import pathlib
f = open('4-30 2percent.txt', 'w')
path='C:/Users/User/Desktop/Git/op3.xlsx'

wb = openpyxl.load_workbook('C:/Users/User/Desktop/Git/op3.xlsx')

sh=wb.active
b=[]

for col in range(2, sh.max_column+1):
    m = round(3.8 + 0.1*col, 1)
    for row in range(1, sh.max_row + 1):
        ener = sh.cell(row=row, column=1).value
        func = sh.cell(row=row, column=col).value
        me = m
        b = [ener, me, func]
        print(*b, sep='   ', file=f)