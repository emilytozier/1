from openpyxl import load_workbook
import operator
wb = load_workbook('trekking1.xlsx')
sheet=wb['Справочник']
calor={}

for i in range(2, sheet.max_row + 1):
    name=sheet.cell(row=i,column=1).value
    bzd = sheet.cell(row=i, column=2).value
    calor[name]=bzd


for k,v in sorted(sorted(calor.items(), key=operator.itemgetter(0)), key=operator.itemgetter(1), reverse=True):
    print(k)

#for i in list_keys:
    #print(i[0])

    #for k, v in sorted(sorted(final_dict.items(), key=operator.itemgetter(0)), key=operator.itemgetter(1),
                       #reverse=True):
        #print(k)