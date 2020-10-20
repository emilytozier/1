<<<<<<< HEAD
#задача чтобы изъять из множества excel все нужные ячейки и вывести их в остортированном виде
=======
#задача чтобы изъять из множества excel файлов (которые в папке rogaikopyta) все нужные ячейки и вывести их в остортированном виде
>>>>>>> e60b6cc6d2d1b82369faf81e00082ae9ec22eebf

import openpyxl
from openpyxl import load_workbook
import os
import pathlib
from collections import OrderedDict

path = 'C:/Users/User/Documents/Visual Studio 2017/DjangoWebProject1/DjangoWebProject1/app/rogaikopyta'

for k in range(1,1001):
    wb = openpyxl.load_workbook('C:/Users/User/Documents/Visual Studio 2017/DjangoWebProject1/DjangoWebProject1/app/rogaikopyta/{id}.xlsx'.format(id=k))
    sheet=wb['Sheet']
    slo={}

    for file in sheet:
        k=sheet.cell(row=2,column=2).value
        z=sheet.cell(row=2,column=4).value
        slo[k]=z
        g=list(slo)
    for i in sorted(g):
        print(i, slo[k])

<<<<<<< HEAD



=======
>>>>>>> e60b6cc6d2d1b82369faf81e00082ae9ec22eebf
