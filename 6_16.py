#Вася составил раскладку по продуктам на один день (она на листе "Раскладка") с указанием названия продукта и его количества в граммах. 
#Посчитайте 4 числа: суммарную калорийность и граммы белков, жиров и углеводов. Числа округлите до целых вниз и введите через пробел.
import openpyxl
import math
wb = openpyxl.load_workbook('trekking2.xlsx')

sheet = wb['Справочник']    
product = dict()
for i in range(2, sheet.max_row+1):
    product[sheet.cell(row=i,column=1).value] = \
        list([sheet.cell(row=i,column=x).value
                      if sheet.cell(row=i,column=x).value != None
                      else 0
                      for x in range(2,sheet.max_column+1)])

sheet = wb['Раскладка']
mday = dict()
for i in range(2, sheet.max_row+1):
    name = sheet.cell(row=i,column=1).value
    massa = sheet.cell(row=i,column=2).value
    mday[name] = mday.get(name, 0) + massa

f_data = [0,0,0,0]
for k,v in mday.items():
    k_data = product[k]
    v_data = [i * (v/100) for i in k_data]
    f_data = [(f_data[i] + v_data[i]) for i in range(len(f_data))]
    
g=list(map(int, f_data))
print(*g)
