
import json

from collections import defaultdict
import openpyxl

wb = openpyxl.load_workbook('data-25290-2019-09-30.xlsx')
fout = open('outa12.json', 'w', encoding='utf-8')

d1 = {}


d2= {}
res = defaultdict(list)
sheet = wb['Sheet0']
for i in range(2, sheet.max_row+1):
    admArea = sheet.cell(row=i,column=5).value
    distr = sheet.cell(row=i,column=6).value
    addr = sheet.cell(row=i,column=7).value
    res[distr].append(addr)

    d1[admArea]=res

json.dump(res, fout, ensure_ascii=False)
fout.close()