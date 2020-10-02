import openpyxl
import math
from collections import Counter
wb = openpyxl.load_workbook('trekking3.xlsx')

sheet = wb['Справочник']
product = dict()
for i in range(2, sheet.max_row + 1):
    product[sheet.cell(row=i, column=1).value] = \
        list([sheet.cell(row=i, column=x).value
              if sheet.cell(row=i, column=x).value != None
              else 0
              for x in range(2, sheet.max_column + 1)])


sheet = wb['Раскладка']
mday = dict()
for i in range(2, sheet.max_row+1):
    name = sheet.cell(row=i,column=2).value
    massa = sheet.cell(row=i,column=3).value
    mday[name]=massa
    #mday[name] = mday.get(name, 0) + massa
    #print(mday)
data=dict()
for i in range(2, sheet.max_row+1):
    day= sheet.cell(row=i,column=1).value
    massa = sheet.cell(row=i, column=3).value
    data[day]=massa
    print(data)

#lo=dict()
#for k, v in mday.items():
    #k_data = product[k]
    #day_data = data[k]
    #for i in range(2,sheet.max_row+1):
        #v_data = [i * (v / 100) for i in k_data] for i in day_data
    #lo[day]=v_data
    #lo[k]=v_data
#print(lo)
    #print(lo)

    #resultdict=dict()
    #for i in range(2,sheet.max_row+1):
        #day = sheet.cell(row=i, column=1).value
        #k_data = product[k]
        #for i in range(2,sheet.max_row+1):
            #chi=v_data
            #resultdict[day]=chi
   #print(resultdict)

#def deposit(arg):
    #name, money = arg
    #bank[name] = bank.setdefault(name, 0) + int(money)
#for _ in range(2, sheet.max_row+1):
    #fun_name = v_data[1:]
    #arg = k_data[0]
    #com[arg](fun_name)
#print(com)



